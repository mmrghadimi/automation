import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_wb(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        c1 = 0.9 * cell.value
        c2 = sheet.cell(row, 4)
        c2.value = c1

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4,
                       )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    filename2 = filename.split('.')[0] + '2.' + filename.split('.')[1]
    wb.save(filename2)


process_wb('file.xlsx')
